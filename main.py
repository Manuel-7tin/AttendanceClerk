# Import required libraries
import io
import re
import json
import time
import string
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl import load_workbook
from streamlit_js_eval import streamlit_js_eval
from openpyxl.utils import range_boundaries
from openpyxl.styles import  Font, Alignment, Border, Side


class Info:
    def __init__(self):
        self.info = {}
        self.error = {}

    def add_info(self, sheet_name, text):
        if self.info.get(sheet_name):
            self.info[sheet_name].append(text)
        else:
            self.info[sheet_name] = [datetime.now().isoformat(), text]
        print("Info now", self.info)


    def add_error(self, sheet_name, text):
        self.error[f"error {len(self.error)}"] = {
            "timestamp": datetime.now().isoformat(),
            "error in": sheet_name,
            "error details": text
        }
        print("Error now", self.error)


    def create_file(self):
        buffer = io.BytesIO()
        for key, value in self.info.items():
            buffer.write(f"{value[0]}|{key}".encode("utf-8"))
            for i, item in enumerate(value):
                buffer.write(item.encode("utf-8") if i != 0 else b"")
                buffer.write(b"\n")
            buffer.write(b"\n\n")
        buffer.seek(0)
        print(self.info)
        print("info content:", buffer.read())
        return buffer


    def create_logfile(self):
        buffer = io.BytesIO()
        buffer.write(json.dumps(self.error).encode("utf-8"))
        buffer.seek(0)
        print(self.error)
        print("JSON content:", buffer.read())
        return buffer


# Note start time
before = datetime.now()
if "stop" not in st.session_state:
    st.session_state.stop = False
if "buffer" not in st.session_state:
    st.session_state.buffer = False
if "info_class" not in st.session_state:
    st.session_state.info_class = Info()
if "time_spent" not in st.session_state:
    st.session_state.time_spent = 0
if "error_message" not in st.session_state:
    st.session_state.error_message = ""
if "file_submitted" not in st.session_state:
    st.session_state.file_submitted = False
if "operation_completed" not in st.session_state:
    st.session_state.operation_completed = False
# placeholder = st.empty()


def match_name(matchee, to_match):
    """
    Determines whether all individual names (words) in `matchee` exist in `to_match`,
    regardless of order or additional words in `to_match`.

    Parameters:
    ----------
    matchee : str
        The name string to be matched. Can be a partial or full name.
    to_match : str
        The target name string that may contain the same name components plus additional ones.

    Returns:
    -------
    bool
        True if all non-empty parts of `matchee` are present in `to_match`, False otherwise.

    Notes:
    -----
    - Matching is case-sensitive by default.
    - Extra words in `to_match` are allowed, but all components in `matchee` must be present.
    - Ignores multiple spaces between words.

    Example:
    --------
    >>> match_name("John Doe", "Mr. John Michael Doe")
    True

    >>> match_name("John Doe", "Doe John")
    True

    >>> match_name("John Doe", "Doe Johnny")
    False
    """
    matchee = [item for item in matchee.split(" ") if item != ""]
    to_match = to_match.split(" ")
    if not matchee:
        return False
    matchee_length, match_size = len(matchee), 0
    for i in range(matchee_length):
        if matchee[i] in to_match:
            to_match.remove(matchee[i])
            match_size += 1
    if match_size == matchee_length:
        return True
    else:
        return False


def caution_split_to_2(text: str, delimiter: str) -> list:
    """
    Splits a string into two parts using the first occurrence of the delimiter.

    If the delimiter appears multiple times, only the first split is used, and the rest
    of the string (including remaining delimiters) is kept as the second part. If the
    delimiter is not found, the second part will be an empty string.

    Args:
        text (str): The input string to be split.
        delimiter (str): The delimiter to split the string on.

    Returns:
        list: A list of two elements - [before_delimiter, after_delimiter].
    """
    first, *rest = text.split(delimiter)
    return [first, delimiter.join(rest)]


def extract_name(name, pp):
    """
    Extracts the actual name from a given string by removing leading codes or prefixes.

    The function is designed to handle input strings that may contain course/session codes
    (e.g., "M25", "J25", or patterns with digits like "M25FRAA") or other prefixes (e.g., "DIP", "DipIFR")
    that precede the real name. It also removes any word matching or closely matching the `pp` parameter
    if it's a short uppercase segment at the start.

    Args:
        name (str): The input string potentially containing a prefix and a name.
        pp (str): A known prefix that may appear in the name and should be removed if present.

    Returns:
        str: The cleaned name with recognized prefixes removed.
    """
    if name == "":
        return ""
    split_name = name.split(" ")
    for i in range(2):
        if split_name[0][:3].lower() == "j25" or split_name[0][:3].lower() == "m25" or split_name[0][1:3].isdigit():
            name = " ".join([item for item in split_name if split_name.index(item) != 0])
        split_name = name.split(" ")
        if split_name[0].isupper() or pp.lower() in split_name[0].lower() and len(split_name[0]) < len(pp) + 3:
            name = " ".join([item for item in split_name if split_name.index(item) != 0])
    if split_name[0].lower() == "dip" or "dipif" in split_name[0].lower():
        name = " ".join([item for item in split_name if split_name.index(item) != 0])
    return name


def safe_force_unmerge(worksheet, cell_range):
    """
    Safely unmerges a range of merged cells in an Excel worksheet using openpyxl.

    This function first ensures that all cells within the specified range exist
    (creating any missing ones), then attempts to unmerge them. This prevents
    `KeyError` or `NoneType` issues that can arise when trying to unmerge
    a range containing non-existent cells.

    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to operate on.
        cell_range (str): The Excel-style range to unmerge (e.g., "A1:G1").

    Returns:
        None
    """
    # Ensure all cells in the range exist before unmerging
    min_col_um, min_row_um, max_col_um, max_row_um = range_boundaries(cell_range)
    for row_um in range(min_row_um, max_row_um + 1):
        for col_um in range(min_col_um, max_col_um + 1):
            _ = worksheet.cell(row=row_um, column=col_um)  # This creates the cell if it doesn't exist

    try:
        worksheet.unmerge_cells(cell_range)
        print(f"Unmerged: {cell_range}")
    except Exception as e:
        print(f"Failed to unmerge {cell_range}: {e}")


def raise_error(e: str, error_type: int, action:str, stop: bool=True):
    if e == st.session_state.error_message:
        return
    st.session_state.error_message = e
    if action == "raise":
        with open(f"error{error_type}xx.html", "r", encoding="utf-8") as f:
            html_err = f.read()
        html_err = html_err.replace("{ERROR_MESSAGE}", st.session_state.get("error_message", "An unknown error occurred."))
        # st.components.v1.html(html_err, scrolling=True, height=700)
        st.session_state.info_class.add_error("raised", html_err)
    elif action[:3] == "log":
        print("logging")
        st.session_state.info_class.add_error(action.split(":")[1], e)
    if stop:
        st.session_state.stop = True
        return True
        # st.stop()  # Prevents further code execution after showing error


def main():
    if not csv_files and excel_file:
        raise_error("File missing", error_type=4, action="raise")
        st.stop()
    # Load workbook
    with st.status("Marking Attendance...", expanded=True) as status:
        st.write("Parsing attendance data...")
        wb = load_workbook(excel_file)
        for j, csv_file in enumerate(csv_files):
            st.session_state.file_submitted = True
            st.session_state.stop = False

            st.write(f"Extracting session {j+1} metadata...")
            time.sleep(0.5)
            # Extract required information about that week's class
            column_names = ["Title", "Details"]
            try:
                class_info = pd.read_csv(csv_file,
                                         sep="\t",
                                         encoding="utf-16",
                                         usecols=[0, 1],
                                         names=column_names,
                                         header=None,
                                         nrows=8)
                paper = class_info["Details"][1].split("(")[0].strip(" ")
                paper = paper.split()[0] + " STANDARD" if "STANDARD" not in paper else paper
                attendees_num = int(class_info["Details"][2])
                day = class_info["Details"][3].split(",")[0]
                date = datetime.strptime(day, "%m/%d/%y")
                print(f"--------------------------___________-----------------------------")

                # Create dataframe containing exclusively information about class attenders
                header_num = 9 if pd.isna(class_info["Title"][7]) else 8
                csv_file.seek(0)
                the_source = pd.read_csv(csv_file, sep="\t", encoding="utf-16", header=header_num)
                cut_off_index = the_source[the_source["Name"] == "3. In-Meeting Activities"].index[0]
                trimmed_source = the_source.iloc[:cut_off_index]
            except Exception as e:
                if raise_error(f"Error extracting metadata: {e}", 4, action="raise"):
                    continue

            with st.spinner(f"Marking {paper} Sheet."):
                # st.write("Selecting required sheet...")
                time.sleep(0.5)
                try:
                    # Load worksheet
                    ws = wb[f"{paper} WK"]

                    # Read worksheet headers directly
                    headers = [cell.value for cell in ws[1]]

                    # Manually convert Excel data into Pandas DataFrame
                    data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        data.append(row)
                    df = pd.DataFrame(data, columns=headers)
                except Exception as e:
                    if raise_error(f"Error parsing main sheet: {e}", 4, "raise"):
                        continue

                # -------------------------------- #
                # Required variables
                # st.write("Validating and recording entries...")
                time.sleep(0.5)
                try:
                    found = []
                    attendees = [name for name in trimmed_source.Name if type(name) == str]
                    print("attendees are: ", attendees)
                    temp_dest = df.copy() # All changes will be done on this copy

                    # Find and mark attendance.
                    for index, row in df.fillna("").iterrows():
                        surname = row["SURNAME"]
                        other_names = row["OTHER NAMES"]
                        # Combine names into a format that matches the weekly attendance report
                        name = f"{surname.strip(" ")} {other_names.strip(" ")}".lower()

                        for attendee in attendees:
                            if match_name(name, attendee.lower()):
                                # This student was found in the list of those who attended the class
                                found.append(attendee)
                                print(name, "found.")
                                attendees.remove(attendee)  # To optimize spped by reducing loop time
                                temp_dest.loc[(temp_dest["SURNAME"].fillna("") == surname) & (
                                            temp_dest["OTHER NAMES"].fillna("") == other_names), date] = "X"
                                break  # To optimize spped by reducing loop time
                except Exception as e:
                    if raise_error(f"Error recording attendees: {e}", 5, "raise"):
                        continue

                # st.write("Resolving discrepancies in attendee logs...")
                time.sleep(0.5)
                try:
                    not_found = [item for item in attendees if item not in found]
                    not_found = list(dict.fromkeys(not_found))  # Remove duplicates
                    # Rename empty column headers/titles to avoid Reindexing error due to non-unique valued Index
                    temp_dest.columns = [
                        col if col is not None else f"unnamed_{i}_srkl12w"
                        for i, col in enumerate(temp_dest.columns)
                    ]
                    # Add attendees that are not in the verified list
                    for name in not_found:
                        if name in ['Teaching Materials', 'Ivy League Manager', "Admin"] or "tutor" in name.lower():
                            # Skip non student attendees
                            continue
                        # Split name into surname first and other names next
                        name = name.replace("(Unverified)", "")
                        name = extract_name(name.replace("(External)", ""), paper.split()[0])
                        names = caution_split_to_2(name, " ")

                        # Create new row containing attendee details
                        new_row = {
                            "S/N": None,  # We'll set this properly below
                            "SURNAME": names[0].title(),
                            "OTHER NAMES": names[1].title(),
                            date: "X"
                        }
                        new_row_df = pd.DataFrame([new_row])

                        # Find the index of the last valid name (excluding the extra info at the end)
                        last_name_index = temp_dest["SURNAME"].last_valid_index()

                        # Detect and separate footer rows
                        footer_rows = temp_dest[~temp_dest["S/N"].astype(str).str.isdigit() & temp_dest["SURNAME"].isna()]
                        main_rows = temp_dest.drop(footer_rows.index).reset_index(drop=True)

                        # Insert new row into main data
                        insert_at = last_name_index + 1
                        main_rows = pd.concat([
                            main_rows.iloc[:insert_at],
                            new_row_df,
                            main_rows.iloc[insert_at:]
                        ], ignore_index=True)

                        # Reassign serial number (S/N)
                        serial = 1
                        for i in range(len(main_rows)):
                            sn = main_rows.at[i, "S/N"]
                            if pd.isna(sn) or isinstance(sn, (int, float)) or (isinstance(sn, str) and sn.strip().isdigit()):
                                main_rows.at[i, "S/N"] = serial
                                serial += 1

                        # Add footer rows back
                        temp_dest = pd.concat([main_rows, footer_rows], ignore_index=True)
                except Exception as e:
                    if raise_error(f"Error sorting missing students: {e}", 5, "raise"):
                        continue

                # Print some stuff
                print(len(found))
                print(not_found)
                # -------------------------------- #

                # st.write("Adjusting layout for footer alignment...")
                time.sleep(0.5)
                try:
                    # Track how many rows you'll insert and where
                    rows_to_insert = len(temp_dest) - ws.max_row + 1
                    insert_at = ws.max_row - 4  # Just above the footer

                    # Determine the number of merged columns in footer rows
                    merge_point = 0
                    if datetime in [type(col) for col in temp_dest.columns]:
                        while type(temp_dest.columns[merge_point]) != datetime:
                            merge_point += 1

                    # Store merged ranges (only for A to the merge_point in the last few rows)
                    original_merges = []
                    for merged_range in ws.merged_cells.ranges:
                        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))  # assuming special rows are at the bottom
                        if min_col == 1 and max_col == merge_point and min_row >= ws.max_row - 5:  # assuming special rows are at the bottom
                            original_merges.append((min_row, max_row))


                    # Insert your new rows
                    if rows_to_insert > 0:
                        ws.insert_rows(idx=insert_at, amount=rows_to_insert)

                    # Figure out how many rows were added and shift accordingly
                    shift = len(temp_dest) - len(df)  # Difference in number of rows after changes
                    # Reapply merged footer cells at their new positions
                    for min_row, max_row in original_merges:
                        ws.merge_cells(start_row=min_row + shift, start_column=1,
                                       end_row=max_row + shift, end_column=merge_point)

                    # Figure out the new footer row location and unmerge merged non footer rows
                    # Note alph[merge_point - 1] gives the column alphabet value of the merge_point
                    new_special_rows = [range_boundaries(str(merged_range))[1] for merged_range in ws.merged_cells.ranges]
                    new_special_rows.sort()
                    alph = string.ascii_uppercase
                    for i in original_merges:
                        if i[0] in new_special_rows[-3:]:
                            continue
                        safe_force_unmerge(ws, f'A{i[0]}:{alph[merge_point - 1]}{i[0]}')
                except Exception as e:
                    if raise_error(f"Merge footer cells error: {e}", 5, f"log:{paper}"):
                        continue

                # st.write("Normalizing font and cell formatting...")
                time.sleep(0.5)
                try:
                    # Correct the alignment and font properties of newly created student entries
                    newly_added_rows = []
                    try:
                        ind = df["SURNAME"].last_valid_index()
                    except ValueError as e:
                        if raise_error(f"Error while identifying new rows: {e}", 5, f"log:{paper}"):
                            continue
                    else:
                        print("ind is", ind)
                        for i, row in temp_dest.iterrows():
                            if i > ind and not pd.isna(row["SURNAME"]):
                                newly_added_rows.append(i + 2)
                    print(f"These are the newly added rows: {newly_added_rows}")


                    # Create a centered alignment and font style
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    custom_font = Font(name=ws["B2"].font.name, size=ws["B2"].font.size)

                    # Knowing `newly_added_rows` is a list of row numbers where new students were added
                    # And Since "S/N" is in column 1 (i.e., column A), change alignment and font
                    for row in newly_added_rows:
                        for col in range(1, merge_point):
                            ws.cell(row=row, column=col).font = custom_font
                            ws.cell(row=row, column=1).alignment = center_alignment
                except Exception as e:
                    if raise_error(f"Error formatting new cells: {e}", 5, f"log:{paper}"):
                        continue

                # st.write("Formula propagation triggered...")
                time.sleep(0.5)
                try:
                    # Ascertain and Update footer rows cell formulas to take the changes into account
                    for i in range(22):
                        if i < merge_point:
                            # Merged cells have no formula
                            continue
                        try:
                            column_label = alph[i] # e.g col H, S, L e.t.c
                            formula_row_start = temp_dest[temp_dest["S/N"] == "TOTAL"].index.item() + 2  # The footer row starting point

                            # This is the cell that contains the COUNTIF formula
                            formula_cell_1 = ws[f'{column_label}{formula_row_start}']
                            formula_1 = formula_cell_1.value # Get the specific formula from the selected cell
                            pattern_1 = rf'({column_label}2:{column_label})(\d+)'  # Regex to determine formula pattern
                            match = re.search(pattern_1, formula_1)
                            if match:
                                # Update formula if it contains a fixed range like H2:H52
                                new_end = formula_row_start - 1
                                new_formula = re.sub(pattern_1, f'{column_label}2:{column_label}{new_end}', formula_1)
                                temp_dest.loc[temp_dest["S/N"] == "TOTAL", [temp_dest.columns[i]]] = new_formula

                            # This is the cell that contains the DIV formula
                            formula_cell_2 = ws[f'{column_label}{formula_row_start + 2}']
                            formula_2 = formula_cell_2.value  # Get the specific formula from the selected cell
                            pattern_2 = rf'({column_label}(\d+)/{column_label})(\d+)'  # Regex to determine formula pattern
                            match = re.search(pattern_2, formula_2)
                            if match:
                                # Update formula if it contains a fixed range like H53/H54
                                new_beg = formula_row_start
                                new_end = formula_row_start + 1
                                new_formula = re.sub(pattern_2, f'{column_label}{new_beg}/{column_label}{new_end}', formula_2)
                                temp_dest.loc[temp_dest["S/N"] == "RATE OF ATTENDANCE PER CLASS", [temp_dest.columns[i]]] = new_formula
                        except TypeError as e:
                            if raise_error(f"Error while updating formulas: {e}", 5, f"log:{paper}", False):
                                continue
                        except Exception as e:
                            if raise_error(f"Error while updating formulas: {e}", 5, "raise", False):
                                continue
                    # Update selected middle footer cell value with number of students in sheet at that time
                    # This number isn't determined by a formula in the Excel sheet so it has to be handwritten
                    temp_dest.loc[temp_dest["S/N"].astype(str).str.strip() == "TOTAL REGISTERED", [date]] = temp_dest["SURNAME"].last_valid_index() + 1
                except TypeError:
                    if raise_error(f"Error updating formulas : {e}", 5, f"log:{paper}"):
                        continue
                except Exception as e:
                    if raise_error(f"Error updating formulas : {e}", 5, "raise"):
                        continue

                # st.write("Saving changes...")
                time.sleep(0.5)
                # Write all changes to the openpyxl worksheet
                for i, row in temp_dest.iterrows():
                    for col in temp_dest.columns:
                        col_idx = temp_dest.columns.get_loc(col) + 1
                        try:
                            ws.cell(row=i + 2, column=col_idx, value=row[col])
                        except AttributeError as e:
                            print("Error while writing to file")
                            print(e)
                            print("Likely trying to write to a read-only merged_cell")
                            print(i+2, col_idx, row[col])
                            if raise_error(f"Error saving changes: {e}", 5, f"log:{paper}", False):
                                continue
                        except ValueError as e:
                            print("Error while writing to file")
                            print(e)
                            print("Likely a ndarray being compared to a single element")
                            print(i+2, col_idx, row[col])
                            if raise_error(f"Error saving changes: {e}", 5, f"log:{paper}", False):
                                continue
                        except Exception as e:
                            if raise_error(f"Error saving changes: {e}", 5, "raise"):
                                continue
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in ws.iter_rows(min_row=insert_at, max_row=rows_to_insert, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border
                st.session_state.info_class.add_info(paper,
                                    f"{attendees_num} people attended the {paper} class held on the {date.strftime("%dth of %B, year %Y")}")
                st.session_state.info_class.add_info(paper,
                                    f"{[man for man in not_found if man in ['Teaching Materials', 'Ivy League Manager', "Admin"] or "tutor" in man.lower()]} were left out of the marking.")
            if st.session_state.stop:
                st.error(f"❌ {paper} marking probably failed.")
            else:
                st.success(f"✅ {paper} marked successfully.")

        st.write("Exporting final document...")
        # Save workbook
        st.session_state.buffer = io.BytesIO()
        wb.save(st.session_state.buffer)
        st.session_state.buffer.seek(0)
        after = datetime.now()
        # wb.save("Bppki.xlsx")
        st.session_state.time_spent = after - before
        st.session_state.operation_completed = True
        status.update(
            label="Marking complete!", state="complete", expanded=False
        )


welcome_text = """<link href="https://fonts.googleapis.com/css2?family=Playfair+Display&display=swap" rel="stylesheet">
<a href="https://github.com/Manuel-7tin/AttendanceClerk" style="text-decoration: none; color: inherit;">
<h1 style='font-family: "Playfair Display", serif; display: inline; font-size: 60px;'>AttendanceClerk</h1>
</a>
<span style='font-size: 17px;'>~ The maestro of your attendance records.</span><br>

##### Powered by precision. Driven by automation."""
st.markdown(welcome_text, unsafe_allow_html=True)
st.write(" ")
col1, col2 = st.columns(2)
with col1:
    excel_file = st.file_uploader("Upload Excel Workbook", type=["xlsx"], disabled=False)
with col2:
    csv_files = st.file_uploader("Upload attendance lists", type=["csv"], disabled=False, accept_multiple_files=True)
st.button(label="Submit files", disabled=st.session_state.file_submitted, on_click=main)

# if csv_file:
#     print("/nFileoo/n", csv_file, excel_file)
if st.session_state.operation_completed:
    total_seconds = st.session_state.time_spent.total_seconds()
    minutes, seconds = int(total_seconds // 60), total_seconds % 60

    # Format as: x minutes and y.yyyyyy seconds
    formatted_time = f"{minutes} minutes and {seconds:.6f} seconds."
    info = "More information in the info-file\nErrors are stored in the log file."
    st.info(f"The total time spent is: {formatted_time}  \n{info}", icon="ℹ️")
    col3, col4 = st.columns(2)
    with col3:
        st.download_button(
            label="Download Output",
            data=st.session_state.buffer,
            file_name="attendance-fixed.xlsx",
            mime="text/csv")
        if st.button("Reload page"):
            streamlit_js_eval(js_expressions="parent.window.location.reload()")
    print("Right before download", st.session_state.info_class.info)
    with col4:
        st.download_button(
            label="Download Info-file",
            data=st.session_state.info_class.create_file(),
            file_name="info-file.txt",
            mime="text/plain")
        st.download_button(
            label="Download LogFile",
            data=st.session_state.info_class.create_logfile(),
            file_name="logs.json",
            mime="application/json")

# Custom HTML & CSS to position buttons
st.markdown(
    """
    <style>
        .corner-button {
            position: fixed;
            top: 60px;
            right: 20px;
            background-color: #0e1117;
            color: white;
            padding: 10px 10px;
            border: 0.5px solid #ffffff;  /* Thin outline */
            cursor: pointer;
            font-size: 16px;
            border-radius: 5px;
        }
        .corner-button:hover {
            background-color: #ffffff;
            color: #0e1117;
            font-weight: bold;
        }
        /* Positioning each button dynamically */
        .corner-buttons a:nth-child(1) button { top: 20px; left: 20px; }  /* Top-left */
        .corner-buttons a:nth-child(2) button { top: 20px; right: 20px; } /* Top-right */
        .corner-buttons a:nth-child(3) button { bottom: 20px; left: 20px; } /* Bottom-left */
        .corner-buttons a:nth-child(4) button { bottom: 20px; right: 20px; } /* Bottom-right */
    </style>
    <a href="https://www.linkedin.com/in/ebi-emmanuel/" target="_blank">
        <button class="corner-button">🔵Linkedin</button>
    </a>
    <a href="https://medium.com/@ebifredrick07" target="_blank">
        <button style="top: 120px;" class="corner-button">✍️Medium</button>
    </a>
    <a href="https://github.com/Manuel-7tin" target="_blank">
        <button style="top: 180px;" class="corner-button">🐙Github</button>
    </a>
    <a href="https://www.imdb.com/title/tt2560140/" target="_blank">
        <button style="top: 240px;" class="corner-button">⚔️AOT</button>
    </a>
    """,
    unsafe_allow_html=True
)